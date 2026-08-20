from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


DOCUMENTS_DIR = Path("documents")
DOCUMENTS_DIR.mkdir(exist_ok=True)


def create_excel_honeytoken(
    token_id: str,
    filename: str = "Staff_Salary_Records.xlsx"
):
    """
    Creates a fake Excel document containing
    a unique honeytoken security callback link.
    """

    filepath = DOCUMENTS_DIR / filename

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Salary Records"

    # ==========================================
    # DOCUMENT TITLE
    # ==========================================

    sheet["A1"] = "STAFF SALARY RECORDS"
    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    # ==========================================
    # TABLE HEADERS
    # ==========================================

    sheet["A3"] = "Employee ID"
    sheet["B3"] = "Department"
    sheet["C3"] = "Position"
    sheet["D3"] = "Monthly Salary"

    # ==========================================
    # FAKE DATA
    # ==========================================

    sample_data = [
        (
            "EMP-001",
            "Administration",
            "Manager",
            "450000"
        ),
        (
            "EMP-002",
            "Finance",
            "Accountant",
            "380000"
        ),
        (
            "EMP-003",
            "IT",
            "Security Analyst",
            "520000"
        ),
        (
            "EMP-004",
            "HR",
            "HR Officer",
            "350000"
        ),
    ]

    for row in sample_data:
        sheet.append(row)

    # ==========================================
    # HONEYTOKEN METADATA
    # ==========================================

    metadata_row = len(sample_data) + 6

    sheet.cell(
        metadata_row,
        1
    ).value = "DOCUMENT SECURITY ID"

    sheet.cell(
        metadata_row,
        2
    ).value = token_id

    sheet.cell(
        metadata_row + 1,
        1
    ).value = "CLASSIFICATION"

    sheet.cell(
        metadata_row + 1,
        2
    ).value = "CONFIDENTIAL"

    # ==========================================
    # SECURITY CALLBACK LINK
    # ==========================================

    callback_url = (
        f"http://127.0.0.1:8000/"
        f"api/events/trigger/{token_id}"
    )

    sheet.cell(
        metadata_row + 3,
        1
    ).value = "SECURITY VERIFICATION"

    sheet.cell(
        metadata_row + 3,
        2
    ).value = "Click here to verify document access"

    sheet.cell(
        metadata_row + 3,
        2
    ).hyperlink = callback_url

    sheet.cell(
        metadata_row + 3,
        2
    ).style = "Hyperlink"

    # ==========================================
    # COLUMN WIDTHS
    # ==========================================

    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 20

    # ==========================================
    # SAVE DOCUMENT
    # ==========================================

    workbook.save(filepath)

    return str(filepath)
